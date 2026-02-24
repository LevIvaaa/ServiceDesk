"""Import charge points from chargePoints (12).xlsx - updates existing and creates new stations."""
import asyncio
import sys
import os

# Add parent dirs to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))

from openpyxl import load_workbook
from sqlalchemy import select
from app.database import async_session_maker
from app.models.operator import Operator
from app.models.station import Station, StationPort


async def main():
    wb = load_workbook('chargePoints (12).xlsx', read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"Total rows in file: {len(rows)}")

    # Collect unique partners
    partners = set()
    for row in rows:
        partner = str(row[13]).strip() if row[13] else None
        if partner:
            partners.add(partner)

    async with async_session_maker() as session:
        # Ensure all operators exist
        operator_map = {}
        for partner_name in sorted(partners):
            result = await session.execute(
                select(Operator).where(Operator.name == partner_name)
            )
            op = result.scalar_one_or_none()
            if not op:
                code = ''.join(c for c in partner_name if c.isalnum())[:15].upper() or "OP"
                # Ensure unique code
                base_code = code
                counter = 1
                while True:
                    check = await session.execute(select(Operator).where(Operator.code == code))
                    if not check.scalar_one_or_none():
                        break
                    code = f"{base_code}{counter}"
                    counter += 1
                op = Operator(name=partner_name, code=code)
                session.add(op)
                await session.flush()
                print(f"  Created operator: {partner_name} ({code})")
            operator_map[partner_name] = op.id
        await session.commit()
        print(f"Operators ready: {len(operator_map)}")

        # Load all existing stations by identifier
        result = await session.execute(select(Station))
        existing = {s.station_id: s for s in result.scalars().all()}
        print(f"Existing stations in DB: {len(existing)}")

        created = 0
        updated = 0
        skipped = 0

        for row in rows:
            identifier = str(row[1]).strip() if row[1] else None
            if not identifier:
                skipped += 1
                continue

            name = str(row[3]).strip() if row[3] else ""
            vendor = str(row[5]).strip() if row[5] else None
            city = str(row[6]).strip() if row[6] else None
            street = str(row[8]).strip() if row[8] else ""
            house = str(row[9]).strip() if row[9] else ""
            lat_raw = row[10]
            lon_raw = row[11]
            partner = str(row[13]).strip() if row[13] else None
            kind = str(row[35]).strip() if row[35] else None
            number = str(row[2]).strip() if row[2] else None

            # Parse lat/lon
            lat = None
            lon = None
            try:
                if lat_raw:
                    lat = float(str(lat_raw).replace(',', '.'))
                if lon_raw:
                    lon = float(str(lon_raw).replace(',', '.'))
            except (ValueError, TypeError):
                pass

            address = f"{street} {house}".strip() or None
            operator_id = operator_map.get(partner) if partner else None
            if not operator_id:
                skipped += 1
                continue

            # Parse connectors
            connectors = []
            for i in range(1, 7):
                ct_idx = 39 + (i - 1) * 5  # connector1type at 39, connector2type at 44, etc.
                cs_idx = ct_idx + 1
                if ct_idx < len(row):
                    ct = str(row[ct_idx]).strip() if row[ct_idx] else None
                    if ct and ct != 'None':
                        cs = str(row[cs_idx]).strip() if cs_idx < len(row) and row[cs_idx] else 'available'
                        connectors.append({'number': i, 'type': ct, 'status': cs})

            if identifier in existing:
                # Update existing station
                station = existing[identifier]
                changed = False
                if vendor and not station.manufacturer:
                    station.manufacturer = vendor
                    changed = True
                if number and not station.station_number:
                    station.station_number = number
                    changed = True
                if city and not station.city:
                    station.city = city
                    changed = True
                if address and not station.address:
                    station.address = address
                    changed = True
                if lat and not station.latitude:
                    station.latitude = lat
                    changed = True
                if lon and not station.longitude:
                    station.longitude = lon
                    changed = True
                if changed:
                    updated += 1
            else:
                # Create new station
                station = Station(
                    station_id=identifier,
                    station_number=number,
                    name=name or identifier,
                    operator_id=operator_id,
                    city=city,
                    address=address,
                    latitude=lat,
                    longitude=lon,
                    manufacturer=vendor,
                    status='online',
                )
                session.add(station)
                await session.flush()

                # Create ports
                for conn in connectors:
                    port = StationPort(
                        station_id=station.id,
                        port_number=conn['number'],
                        connector_type=conn['type'],
                        status=conn['status'] if conn['status'] != 'None' else 'available',
                    )
                    session.add(port)

                existing[identifier] = station
                created += 1

            if (created + updated) % 500 == 0 and (created + updated) > 0:
                print(f"  Progress: created={created}, updated={updated}")

        await session.commit()
        print(f"\nDone! Created: {created}, Updated: {updated}, Skipped: {skipped}")

    wb.close()


if __name__ == "__main__":
    asyncio.run(main())
