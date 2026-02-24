import json
import os
import uuid
from datetime import datetime
from pathlib import Path
from typing import Annotated, Optional

import aiofiles
from pydantic import BaseModel as PydanticBaseModel
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from sqlalchemy import func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.api.deps import CurrentUser, DbSession, PermissionRequired
from app.config import settings
from app.models.department import Department
from app.models.operator import Operator
from app.models.station import Station
from app.models.ticket import Ticket, TicketAttachment, TicketComment, TicketHistory, TicketLog
from app.models.user import User
from app.schemas.common import PaginatedResponse
from app.schemas.ticket import (
    ParseMessageRequest,
    ParseMessageResponse,
    TicketAssignUpdate,
    TicketAttachmentResponse,
    TicketCommentCreate,
    TicketCommentResponse,
    TicketCreate,
    TicketDelegateUpdate,
    TicketDetailResponse,
    TicketHistoryResponse,
    TicketListResponse,
    TicketLogResponse,
    TicketResponse,
    TicketStatusUpdate,
    TicketUpdate,
)
from app.services.ticket_service import TicketService
from app.services.assignment_service import AssignmentService
from app.services.notification_service import NotificationService

router = APIRouter()


async def generate_ticket_number(db: AsyncSession) -> str:
    """Generate unique ticket number: EF-YYYY-NNNNNN with retry logic"""
    year = datetime.utcnow().year
    max_retries = 5
    
    for attempt in range(max_retries):
        # Get the maximum ticket number for this year
        result = await db.execute(
            select(func.max(Ticket.ticket_number))
            .where(Ticket.ticket_number.like(f"EF-{year}-%"))
        )
        max_number = result.scalar()
        
        if max_number:
            # Extract counter from last ticket number
            try:
                counter = int(max_number.split('-')[-1]) + 1
            except (ValueError, IndexError):
                counter = 1
        else:
            counter = 1
        
        ticket_number = f"EF-{year}-{counter:06d}"
        
        # Check if this number already exists
        check_result = await db.execute(
            select(Ticket.id).where(Ticket.ticket_number == ticket_number)
        )
        if not check_result.scalar():
            return ticket_number
        
        # If exists, try next number
        counter += 1
    
    # Fallback: use timestamp-based unique number
    import time
    timestamp = int(time.time() * 1000) % 1000000
    return f"EF-{year}-{timestamp:06d}"


@router.get("", response_model=PaginatedResponse[TicketListResponse])
async def list_tickets(
    db: DbSession,
    current_user: CurrentUser,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    search: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    category: Optional[str] = None,
    assigned_user_id: Optional[int] = None,
    assigned_department_id: Optional[int] = None,
    department_id: Optional[int] = None,
    station_id: Optional[int] = None,
    created_by_id: Optional[int] = None,
    my_tickets: bool = False,
    delegated_to_me: bool = False,
):
    """List tickets with pagination and filters."""
    query = select(Ticket).options(
        selectinload(Ticket.station).selectinload(Station.operator),
        selectinload(Ticket.assigned_user),
        selectinload(Ticket.assigned_department),
        selectinload(Ticket.created_by),
    )

    # Apply filters
    if search:
        search_filter = f"%{search}%"
        query = query.outerjoin(Station, Ticket.station_id == Station.id).where(
            or_(
                Ticket.ticket_number.ilike(search_filter),
                Ticket.title.ilike(search_filter),
                Station.station_number.ilike(search_filter),
                Station.station_id.ilike(search_filter),
            )
        )
    if status:
        # Support multiple statuses separated by comma
        if ',' in status:
            statuses = [s.strip() for s in status.split(',')]
            query = query.where(Ticket.status.in_(statuses))
        else:
            query = query.where(Ticket.status == status)
    if priority:
        query = query.where(Ticket.priority == priority)
    if category:
        query = query.where(Ticket.category == category)
    if assigned_user_id is not None:
        query = query.where(Ticket.assigned_user_id == assigned_user_id)
    if assigned_department_id is not None:
        query = query.where(Ticket.assigned_department_id == assigned_department_id)
    if department_id is not None:
        # Filter by department (either assigned or related)
        query = query.where(Ticket.assigned_department_id == department_id)
    if station_id is not None:
        query = query.where(Ticket.station_id == station_id)
    if created_by_id is not None:
        query = query.where(Ticket.created_by_id == created_by_id)
    if my_tickets:
        query = query.where(
            or_(
                Ticket.assigned_user_id == current_user.id,
                Ticket.created_by_id == current_user.id,
            )
        )
    if delegated_to_me:
        # Only tickets delegated specifically to the current user
        delegated_subq = (
            select(TicketHistory.ticket_id)
            .where(TicketHistory.action == "delegated")
            .distinct()
        )
        query = query.where(
            Ticket.id.in_(delegated_subq),
            Ticket.assigned_user_id == current_user.id,
        )

    # Count total
    count_query = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_query)).scalar()

    # Apply pagination
    offset = (page - 1) * per_page
    query = query.offset(offset).limit(per_page).order_by(Ticket.created_at.desc())

    result = await db.execute(query)
    tickets = result.scalars().all()

    # Build response with counts
    items = []
    for ticket in tickets:
        ticket_dict = TicketListResponse.model_validate(ticket).model_dump()
        if ticket.station:
            ticket_dict["station"] = {
                "id": ticket.station.id,
                "station_id": ticket.station.station_id,
                "station_number": ticket.station.station_number,
                "name": ticket.station.name,
                "address": ticket.station.address,
                "operator_name": ticket.station.operator.name if ticket.station.operator else "",
            }
        items.append(TicketListResponse(**ticket_dict))

    return PaginatedResponse(
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        pages=(total + per_page - 1) // per_page,
    )


@router.post("", response_model=TicketResponse, status_code=status.HTTP_201_CREATED)
async def create_ticket(
    ticket_data: TicketCreate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.create"))],
):
    """Create a new ticket."""
    ticket_number = await generate_ticket_number(db)

    ticket = Ticket(
        ticket_number=ticket_number,
        created_by_id=current_user.id,
        **ticket_data.model_dump(),
    )
    db.add(ticket)
    await db.flush()

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="created",
        new_value=json.dumps({"ticket_number": ticket_number}),
    )
    db.add(history)

    await db.commit()

    # Reload ticket with all relationships properly loaded (including Station.operator)
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket.id)
    )
    ticket = result.scalar_one()

    # Send notification about new ticket
    from app.services.notification_service import NotificationService
    notification_service = NotificationService(db)
    await notification_service.notify_ticket_created(ticket)

    return await _build_ticket_response(ticket, db)


@router.get("/export")
async def export_tickets(
    db: DbSession,
    current_user: CurrentUser,
):
    """Export tickets to Excel file with all ticket data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        query = (
            select(Ticket)
            .options(
                selectinload(Ticket.station).selectinload(Station.operator),
                selectinload(Ticket.assigned_user),
                selectinload(Ticket.assigned_department),
                selectinload(Ticket.created_by),
            )
            .order_by(Ticket.created_at.desc())
        )

        result = await db.execute(query)
        tickets = result.scalars().all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"

        headers = [
            "Номер тікету", "Заголовок", "Опис", "Категорія", "Пріоритет", "Статус",
            "Тип інциденту", "Тип клієнта", "Джерело звернення",
            "Номер станції", "Назва станції", "ID станції", "Адреса станції", "Власник станції",
            "Тип порту", "Модель авто",
            "Ім'я заявника", "Телефон заявника", "Email заявника",
            "Відповідальний", "Відділ", "Створив",
            "Дата створення", "Дата оновлення", "Дата вирішення", "Дата закриття",
            "SLA дедлайн", "SLA порушено", "Логи станції",
        ]

        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_num, ticket in enumerate(tickets, 2):
            station_number = ticket.station.station_number if ticket.station else ""
            station_name = ticket.station.name if ticket.station else ""
            station_id = ticket.station.station_id if ticket.station else ""
            station_address = ticket.station.address if ticket.station else ""
            station_operator = ticket.station.operator.name if ticket.station and ticket.station.operator else ""
            
            row_data = [
                ticket.ticket_number, ticket.title, ticket.description,
                ticket.category, ticket.priority, ticket.status,
                ticket.incident_type or "", ticket.client_type or "", ticket.contact_source or "",
                station_number, station_name, station_id, station_address, station_operator,
                ticket.port_type or "", ticket.vehicle or "",
                ticket.reporter_name or "", ticket.reporter_phone or "", ticket.reporter_email or "",
                f"{ticket.assigned_user.first_name} {ticket.assigned_user.last_name}" if ticket.assigned_user else "",
                ticket.assigned_department.name if ticket.assigned_department else "",
                f"{ticket.created_by.first_name} {ticket.created_by.last_name}" if ticket.created_by else "",
                ticket.created_at.strftime("%d.%m.%Y %H:%M") if ticket.created_at else "",
                ticket.updated_at.strftime("%d.%m.%Y %H:%M") if ticket.updated_at else "",
                ticket.resolved_at.strftime("%d.%m.%Y %H:%M") if ticket.resolved_at else "",
                ticket.closed_at.strftime("%d.%m.%Y %H:%M") if ticket.closed_at else "",
                ticket.sla_due_date.strftime("%d.%m.%Y %H:%M") if ticket.sla_due_date else "",
                "Так" if ticket.sla_breached else "Ні",
                ticket.station_logs or "",
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from fastapi.responses import StreamingResponse
        filename = f"tickets_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Export failed: {str(e)}"
        )


@router.get("/{ticket_id}", response_model=TicketDetailResponse)
async def get_ticket(
    ticket_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get a specific ticket with all details."""
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
            selectinload(Ticket.comments).selectinload(TicketComment.user),
            selectinload(Ticket.attachments).selectinload(TicketAttachment.uploaded_by),
            selectinload(Ticket.history).selectinload(TicketHistory.user),
            selectinload(Ticket.logs),
        )
        .where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    return await _build_ticket_detail_response(ticket, db, current_user)


@router.put("/{ticket_id}", response_model=TicketResponse)
async def update_ticket(
    ticket_id: int,
    ticket_data: TicketUpdate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.edit"))],
):
    """Update a ticket."""
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Track changes for history
    old_values = {}
    new_values = {}
    update_data = ticket_data.model_dump(exclude_unset=True)

    for field, value in update_data.items():
        old_value = getattr(ticket, field)
        if old_value != value:
            old_values[field] = old_value
            new_values[field] = value
        setattr(ticket, field, value)

    if old_values:
        history = TicketHistory(
            ticket_id=ticket.id,
            user_id=current_user.id,
            action="updated",
            old_value=json.dumps(old_values, default=str),
            new_value=json.dumps(new_values, default=str),
        )
        db.add(history)

    await db.commit()

    # Reload ticket with all relationships properly loaded (including Station.operator)
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket.id)
    )
    ticket = result.scalar_one()

    return await _build_ticket_response(ticket, db)


@router.delete("/{ticket_id}")
async def delete_ticket(
    ticket_id: int,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.delete"))],
):
    """Delete a ticket. Only closed tickets can be deleted."""
    result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Admins can delete any ticket; others only new/closed
    if not current_user.is_admin:
        if ticket.status not in ["new", "closed"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only new or closed tickets can be deleted",
            )

    await db.delete(ticket)
    await db.commit()

    return {"message": "Ticket deleted successfully"}


@router.put("/{ticket_id}/status", response_model=TicketResponse)
async def update_ticket_status(
    ticket_id: int,
    status_data: TicketStatusUpdate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.change_status"))],
):
    """Update ticket status."""
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Перевірка: звичайні користувачі можуть закривати будь-які тікети
    # (раніше було обмеження тільки на свої)

    old_status = ticket.status
    ticket.status = status_data.status

    # Set timestamps
    if status_data.status == "reviewing":
        ticket.resolved_at = datetime.utcnow()
    elif status_data.status == "closed":
        ticket.closed_at = datetime.utcnow()

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="status_changed",
        old_value=json.dumps({"status": old_status}),
        new_value=json.dumps({"status": status_data.status}),
    )
    db.add(history)

    # Add comment if provided
    if status_data.comment:
        comment = TicketComment(
            ticket_id=ticket.id,
            user_id=current_user.id,
            content=status_data.comment,
            is_internal=True,
        )
        db.add(comment)

    await db.commit()

    # Send notification on any status change
    if old_status != status_data.status:
        from app.services.notification_service import NotificationService
        notification_service = NotificationService(db)
        await notification_service.notify_ticket_status_changed(
            ticket, old_status, status_data.status, current_user
        )

    # Reload ticket with all relationships properly loaded (including Station.operator)
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket.id)
    )
    ticket = result.scalar_one()

    return await _build_ticket_response(ticket, db)


@router.put("/{ticket_id}/assign", response_model=TicketResponse)
async def assign_ticket(
    ticket_id: int,
    assign_data: TicketAssignUpdate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.assign"))],
):
    """Assign ticket to a user."""
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    old_assigned = ticket.assigned_user_id
    ticket.assigned_user_id = assign_data.assigned_user_id

    # Update status if new
    if ticket.status == "new":
        ticket.status = "in_progress"

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="assigned",
        old_value=json.dumps({"assigned_user_id": old_assigned}),
        new_value=json.dumps({"assigned_user_id": assign_data.assigned_user_id}),
    )
    db.add(history)

    # Add comment if provided
    if assign_data.comment:
        comment = TicketComment(
            ticket_id=ticket.id,
            user_id=current_user.id,
            content=assign_data.comment,
            is_internal=False,
        )
        db.add(comment)

    await db.commit()

    # Reload ticket with all relationships properly loaded (including Station.operator)
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket.id)
    )
    ticket = result.scalar_one()

    return await _build_ticket_response(ticket, db)


@router.put("/{ticket_id}/delegate", response_model=TicketResponse)
async def delegate_ticket(
    ticket_id: int,
    delegate_data: TicketDelegateUpdate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.delegate"))],
):
    """Delegate ticket to another department with optional auto-assignment."""
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket_id)
    )
    ticket = result.scalar_one_or_none()

    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Verify department exists
    dept_result = await db.execute(
        select(Department).where(Department.id == delegate_data.assigned_department_id)
    )
    if not dept_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Department not found",
        )

    old_dept = ticket.assigned_department_id
    old_user = ticket.assigned_user_id
    old_status = ticket.status

    ticket.assigned_department_id = delegate_data.assigned_department_id

    # Assign user only if explicitly specified
    # If not specified, leave it empty so someone from the department can accept it
    assigned_user_id = delegate_data.assigned_user_id
    ticket.assigned_user_id = assigned_user_id

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket.id,
        user_id=current_user.id,
        action="delegated",
        old_value=json.dumps({
            "assigned_department_id": old_dept,
            "assigned_user_id": old_user,
        }),
        new_value=json.dumps({
            "assigned_department_id": delegate_data.assigned_department_id,
            "assigned_user_id": assigned_user_id,
        }),
    )
    db.add(history)

    # Add comment if provided
    if delegate_data.comment:
        comment = TicketComment(
            ticket_id=ticket.id,
            user_id=current_user.id,
            content=delegate_data.comment,
            is_internal=False,
        )
        db.add(comment)

    await db.commit()

    # Send notification to assigned user
    if assigned_user_id and assigned_user_id != old_user:
        notification_service = NotificationService(db)
        assigned_user = await db.get(User, assigned_user_id)
        if assigned_user:
            # Reload ticket for notification
            result = await db.execute(
                select(Ticket)
                .options(selectinload(Ticket.created_by))
                .where(Ticket.id == ticket.id)
            )
            ticket_for_notif = result.scalar_one()
            await notification_service.notify_ticket_assigned(ticket_for_notif, assigned_user)

    # Notify new department about delegated ticket
    if old_dept != delegate_data.assigned_department_id:
        notification_service = NotificationService(db)
        # Reload ticket for notification
        result = await db.execute(
            select(Ticket)
            .options(selectinload(Ticket.created_by), selectinload(Ticket.assigned_department))
            .where(Ticket.id == ticket.id)
        )
        ticket_for_notif = result.scalar_one()
        await notification_service.notify_ticket_created(ticket_for_notif)

    # Reload ticket with all relationships properly loaded (including Station.operator)
    result = await db.execute(
        select(Ticket)
        .options(
            selectinload(Ticket.station).selectinload(Station.operator),
            selectinload(Ticket.assigned_user),
            selectinload(Ticket.assigned_department),
            selectinload(Ticket.created_by),
        )
        .where(Ticket.id == ticket.id)
    )
    ticket = result.scalar_one()

    return await _build_ticket_response(ticket, db)


@router.post("/{ticket_id}/comments", response_model=TicketCommentResponse, status_code=status.HTTP_201_CREATED)
async def add_comment(
    ticket_id: int,
    comment_data: TicketCommentCreate,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.add_comment"))],
):
    """Add a comment to a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(
        select(Ticket)
        .options(selectinload(Ticket.created_by))
        .where(Ticket.id == ticket_id)
    )
    ticket = ticket_result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    comment = TicketComment(
        ticket_id=ticket_id,
        user_id=current_user.id,
        content=comment_data.content,
        is_internal=comment_data.is_internal,
    )
    db.add(comment)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="commented",
        new_value=json.dumps({"is_internal": comment_data.is_internal}),
    )
    db.add(history)

    await db.commit()
    await db.refresh(comment, ["user"])

    # Send notification about new comment
    notification_service = NotificationService(db)
    await notification_service.notify_ticket_commented(ticket, current_user, comment_data.is_internal)

    return TicketCommentResponse.model_validate(comment)


@router.get("/{ticket_id}/comments", response_model=list[TicketCommentResponse])
async def get_comments(
    ticket_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get all comments for a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Check if user can see internal comments
    from app.core.permissions import check_permission
    can_view_internal = await check_permission(
        current_user, "tickets.view_internal_comments", db
    )

    query = (
        select(TicketComment)
        .options(selectinload(TicketComment.user))
        .where(TicketComment.ticket_id == ticket_id)
    )

    if not can_view_internal:
        query = query.where(TicketComment.is_internal == False)

    query = query.order_by(TicketComment.created_at)

    result = await db.execute(query)
    comments = result.scalars().all()

    return [TicketCommentResponse.model_validate(c) for c in comments]


@router.get("/{ticket_id}/history", response_model=list[TicketHistoryResponse])
async def get_history(
    ticket_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get ticket history."""
    # Verify ticket exists
    ticket_result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    result = await db.execute(
        select(TicketHistory)
        .options(selectinload(TicketHistory.user))
        .where(TicketHistory.ticket_id == ticket_id)
        .order_by(TicketHistory.created_at.desc())
    )
    history = result.scalars().all()

    return [TicketHistoryResponse.model_validate(h) for h in history]


async def _build_ticket_response(ticket: Ticket, db: AsyncSession) -> TicketResponse:
    """Build ticket response with counts."""
    # Count comments
    comments_count = (await db.execute(
        select(func.count()).where(TicketComment.ticket_id == ticket.id)
    )).scalar()

    # Count attachments
    attachments_count = (await db.execute(
        select(func.count()).where(TicketAttachment.ticket_id == ticket.id)
    )).scalar()

    response_dict = {
        "id": ticket.id,
        "ticket_number": ticket.ticket_number,
        "title": ticket.title,
        "description": ticket.description,
        "category": ticket.category,
        "priority": ticket.priority,
        "status": ticket.status,
        "station_id": ticket.station_id,
        "station": None,
        "port_number": ticket.port_number,
        "reporter_name": ticket.reporter_name,
        "reporter_phone": ticket.reporter_phone,
        "reporter_email": ticket.reporter_email,
        "assigned_user_id": ticket.assigned_user_id,
        "assigned_user": ticket.assigned_user,
        "assigned_department_id": ticket.assigned_department_id,
        "assigned_department": ticket.assigned_department,
        "created_by_id": ticket.created_by_id,
        "created_by": ticket.created_by,
        "created_at": ticket.created_at,
        "updated_at": ticket.updated_at,
        "resolved_at": ticket.resolved_at,
        "closed_at": ticket.closed_at,
        "sla_due_date": ticket.sla_due_date,
        "sla_breached": ticket.sla_breached,
        "ai_log_analysis": ticket.ai_log_analysis,
        "comments_count": comments_count,
        "attachments_count": attachments_count,
        # New fields from TZ
        "incident_type": ticket.incident_type,
        "port_type": ticket.port_type,
        "contact_source": ticket.contact_source,
        "station_logs": ticket.station_logs,
        "vehicle": ticket.vehicle,
        "client_type": ticket.client_type,
    }

    if ticket.station:
        response_dict["station"] = {
            "id": ticket.station.id,
            "station_id": ticket.station.station_id,
            "station_number": ticket.station.station_number,
            "name": ticket.station.name,
            "address": ticket.station.address,
            "operator_name": ticket.station.operator.name if ticket.station.operator else "",
        }

    return TicketResponse(**response_dict)


async def _build_ticket_detail_response(
    ticket: Ticket, db: AsyncSession, current_user: User
) -> TicketDetailResponse:
    """Build detailed ticket response."""
    from app.core.permissions import check_permission

    can_view_internal = await check_permission(
        current_user, "tickets.view_internal_comments", db
    )

    # Filter internal comments if needed
    comments = ticket.comments
    if not can_view_internal:
        comments = [c for c in comments if not c.is_internal]

    response_dict = {
        "id": ticket.id,
        "ticket_number": ticket.ticket_number,
        "title": ticket.title,
        "description": ticket.description,
        "category": ticket.category,
        "priority": ticket.priority,
        "status": ticket.status,
        "station_id": ticket.station_id,
        "station": None,
        "port_number": ticket.port_number,
        "reporter_name": ticket.reporter_name,
        "reporter_phone": ticket.reporter_phone,
        "reporter_email": ticket.reporter_email,
        "assigned_user_id": ticket.assigned_user_id,
        "assigned_user": ticket.assigned_user,
        "assigned_department_id": ticket.assigned_department_id,
        "assigned_department": ticket.assigned_department,
        "created_by_id": ticket.created_by_id,
        "created_by": ticket.created_by,
        "created_at": ticket.created_at,
        "updated_at": ticket.updated_at,
        "resolved_at": ticket.resolved_at,
        "closed_at": ticket.closed_at,
        "sla_due_date": ticket.sla_due_date,
        "sla_breached": ticket.sla_breached,
        "ai_log_analysis": ticket.ai_log_analysis,
        "comments_count": len(comments),
        "attachments_count": len(ticket.attachments),
        "comments": comments,
        "attachments": ticket.attachments,
        "history": sorted(ticket.history, key=lambda h: h.created_at, reverse=True),
        "logs": ticket.logs,
        # New fields from TZ
        "incident_type": ticket.incident_type,
        "port_type": ticket.port_type,
        "contact_source": ticket.contact_source,
        "station_logs": ticket.station_logs,
        "vehicle": ticket.vehicle,
        "client_type": ticket.client_type,
    }

    if ticket.station:
        response_dict["station"] = {
            "id": ticket.station.id,
            "station_id": ticket.station.station_id,
            "station_number": ticket.station.station_number,
            "name": ticket.station.name,
            "address": ticket.station.address,
            "operator_name": ticket.station.operator.name if ticket.station.operator else "",
        }
        response_dict["station"] = {
            "id": ticket.station.id,
            "station_id": ticket.station.station_id,
            "name": ticket.station.name,
            "address": ticket.station.address,
            "operator_name": ticket.station.operator.name if ticket.station.operator else "",
        }

    return TicketDetailResponse(**response_dict)


# ============== Logs endpoints ==============

@router.post("/{ticket_id}/logs", response_model=TicketLogResponse, status_code=status.HTTP_201_CREATED)
async def upload_ticket_log(
    ticket_id: int,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.add_logs"))],
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    log_start_time: Optional[datetime] = Form(None),
    log_end_time: Optional[datetime] = Form(None),
):
    """Upload a log file to a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(
        select(Ticket).where(Ticket.id == ticket_id)
    )
    ticket = ticket_result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Check file size
    content = await file.read()
    if len(content) > settings.MAX_UPLOAD_SIZE:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"File too large. Maximum size is {settings.MAX_UPLOAD_SIZE // (1024*1024)}MB",
        )

    # Create storage directory
    storage_path = Path(settings.LOGS_STORAGE_PATH) / str(ticket_id)
    storage_path.mkdir(parents=True, exist_ok=True)

    # Generate unique filename
    file_ext = Path(file.filename).suffix if file.filename else ".log"
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = storage_path / unique_filename

    # Save file
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    # Create log record
    log = TicketLog(
        ticket_id=ticket_id,
        log_type="manual",
        filename=file.filename or unique_filename,
        file_path=str(file_path),
        file_size=len(content),
        station_id=ticket.station_id,
        log_start_time=log_start_time,
        log_end_time=log_end_time,
        description=description,
    )
    db.add(log)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="log_uploaded",
        new_value=json.dumps({"filename": file.filename}),
    )
    db.add(history)

    await db.commit()
    await db.refresh(log)

    return TicketLogResponse.model_validate(log)


class TextLogRequest(PydanticBaseModel):
    content: str
    description: Optional[str] = None


@router.post("/{ticket_id}/logs/text", response_model=TicketLogResponse, status_code=status.HTTP_201_CREATED)
async def upload_text_log(
    ticket_id: int,
    request: TextLogRequest,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.add_logs"))],
):
    """Upload a text log to a ticket (for pasting OCPP logs, etc.)."""
    # Verify ticket exists
    ticket_result = await db.execute(
        select(Ticket).where(Ticket.id == ticket_id)
    )
    ticket = ticket_result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Create storage directory
    storage_path = Path(settings.LOGS_STORAGE_PATH) / str(ticket_id)
    storage_path.mkdir(parents=True, exist_ok=True)

    # Generate unique filename
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    unique_filename = f"text_log_{timestamp}_{uuid.uuid4().hex[:8]}.log"
    file_path = storage_path / unique_filename

    # Save content to file
    content_bytes = request.content.encode("utf-8")
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content_bytes)

    # Create log record
    log = TicketLog(
        ticket_id=ticket_id,
        log_type="text",
        filename=unique_filename,
        file_path=str(file_path),
        file_size=len(content_bytes),
        station_id=ticket.station_id,
        description=request.description or "Pasted text log",
    )
    db.add(log)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="log_uploaded",
        new_value=json.dumps({"filename": unique_filename, "type": "text"}),
    )
    db.add(history)

    await db.commit()
    await db.refresh(log)

    return TicketLogResponse.model_validate(log)


@router.get("/{ticket_id}/logs", response_model=list[TicketLogResponse])
async def get_ticket_logs(
    ticket_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get all logs for a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    result = await db.execute(
        select(TicketLog)
        .where(TicketLog.ticket_id == ticket_id)
        .order_by(TicketLog.collected_at.desc())
    )
    logs = result.scalars().all()

    return [TicketLogResponse.model_validate(log) for log in logs]


@router.get("/{ticket_id}/logs/{log_id}/download")
async def download_ticket_log(
    ticket_id: int,
    log_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Download a log file."""
    result = await db.execute(
        select(TicketLog).where(
            TicketLog.id == log_id,
            TicketLog.ticket_id == ticket_id,
        )
    )
    log = result.scalar_one_or_none()

    if not log:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Log not found",
        )

    if not os.path.exists(log.file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Log file not found on disk",
        )

    return FileResponse(
        path=log.file_path,
        filename=log.filename,
        media_type="application/octet-stream",
    )


@router.delete("/{ticket_id}/logs/{log_id}")
async def delete_ticket_log(
    ticket_id: int,
    log_id: int,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.delete_logs"))],
):
    """Delete a log file."""
    result = await db.execute(
        select(TicketLog).where(
            TicketLog.id == log_id,
            TicketLog.ticket_id == ticket_id,
        )
    )
    log = result.scalar_one_or_none()

    if not log:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Log not found",
        )

    # Delete file from disk
    if os.path.exists(log.file_path):
        os.remove(log.file_path)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="log_deleted",
        old_value=json.dumps({"filename": log.filename}),
    )
    db.add(history)

    await db.delete(log)
    await db.commit()

    return {"message": "Log deleted successfully"}


# ============== Message parsing endpoint ==============

@router.post("/parse-message", response_model=ParseMessageResponse)
async def parse_customer_message(
    request: ParseMessageRequest,
    db: DbSession,
    current_user: CurrentUser,
):
    """Parse customer message using LLM to extract ticket data."""
    from app.services.message_parser_service import message_parser

    # Parse the message
    parsed = await message_parser.parse_message(request.message)

    # Try to find station in database if station_id was extracted
    station_db_id = None
    station_found = False
    if parsed.station_id:
        # Search by station_id or external_id (operator's number)
        result = await db.execute(
            select(Station).where(
                or_(
                    Station.station_id.ilike(f"%{parsed.station_id}%"),
                    Station.external_id.ilike(f"%{parsed.station_id}%")
                )
            ).limit(1)
        )
        station = result.scalar_one_or_none()
        if station:
            station_db_id = station.id
            station_found = True

    # Try to find operator in database if operator_name was extracted
    operator_db_id = None
    operator_found = False
    if parsed.operator_name:
        # Search by name (case-insensitive, partial match) - take first match
        result = await db.execute(
            select(Operator).where(
                or_(
                    Operator.name.ilike(f"%{parsed.operator_name}%"),
                    Operator.code.ilike(f"%{parsed.operator_name}%")
                )
            ).limit(1)
        )
        operator = result.scalar_one_or_none()
        if operator:
            operator_db_id = operator.id
            operator_found = True

    return ParseMessageResponse(
        title=parsed.title,
        description=parsed.description,
        category=parsed.category,
        priority=parsed.priority,
        # Station info
        station_id=parsed.station_id,
        station_db_id=station_db_id,
        station_name=parsed.station_name,
        station_address=parsed.station_address,
        station_city=parsed.station_city,
        station_found=station_found,
        # Operator info
        operator_name=parsed.operator_name,
        operator_db_id=operator_db_id,
        operator_found=operator_found,
        # Port and vehicle
        port_number=parsed.port_number,
        vehicle_info=parsed.vehicle_info,
        # Reporter info
        reporter_name=parsed.reporter_name,
        reporter_phone=parsed.reporter_phone,
        reporter_email=parsed.reporter_email,
    )


# ============== Attachments endpoints ==============

@router.get("/{ticket_id}/attachments", response_model=list[TicketAttachmentResponse])
async def get_ticket_attachments(
    ticket_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Get all attachments for a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    if not ticket_result.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    result = await db.execute(
        select(TicketAttachment)
        .options(selectinload(TicketAttachment.uploaded_by))
        .where(TicketAttachment.ticket_id == ticket_id)
        .order_by(TicketAttachment.uploaded_at.desc())
    )
    attachments = result.scalars().all()

    return [TicketAttachmentResponse.model_validate(a) for a in attachments]


@router.post("/{ticket_id}/attachments", response_model=TicketAttachmentResponse, status_code=status.HTTP_201_CREATED)
async def upload_ticket_attachment(
    ticket_id: int,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.upload_attachments"))],
    file: UploadFile = File(...),
):
    """Upload an attachment to a ticket."""
    # Verify ticket exists
    ticket_result = await db.execute(select(Ticket).where(Ticket.id == ticket_id))
    ticket = ticket_result.scalar_one_or_none()
    if not ticket:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Ticket not found",
        )

    # Check file size (100MB)
    content = await file.read()
    max_size = 100 * 1024 * 1024  # 100MB
    if len(content) > max_size:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail="File too large. Maximum size is 100MB",
        )

    # Create storage directory
    storage_path = Path(settings.ATTACHMENTS_STORAGE_PATH) / str(ticket_id)
    storage_path.mkdir(parents=True, exist_ok=True)

    # Generate unique filename
    file_ext = Path(file.filename).suffix if file.filename else ".bin"
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = storage_path / unique_filename

    # Save file
    async with aiofiles.open(file_path, "wb") as f:
        await f.write(content)

    # Determine mime type - fallback to extension-based detection
    import mimetypes
    mime = file.content_type
    if not mime or mime == "application/octet-stream":
        guessed, _ = mimetypes.guess_type(file.filename or "")
        if guessed:
            mime = guessed

    # Create attachment record
    attachment = TicketAttachment(
        ticket_id=ticket_id,
        filename=file.filename or unique_filename,
        file_path=str(file_path),
        file_size=len(content),
        mime_type=mime or "application/octet-stream",
        uploaded_by_id=current_user.id,
    )
    db.add(attachment)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="attachment_uploaded",
        new_value=json.dumps({"filename": file.filename}),
    )
    db.add(history)

    await db.commit()
    await db.refresh(attachment, ["uploaded_by"])

    return TicketAttachmentResponse.model_validate(attachment)


@router.get("/{ticket_id}/attachments/{attachment_id}/download")
async def download_ticket_attachment(
    ticket_id: int,
    attachment_id: int,
    db: DbSession,
    current_user: CurrentUser,
):
    """Download an attachment file."""
    result = await db.execute(
        select(TicketAttachment).where(
            TicketAttachment.id == attachment_id,
            TicketAttachment.ticket_id == ticket_id,
        )
    )
    attachment = result.scalar_one_or_none()

    if not attachment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attachment not found",
        )

    if not os.path.exists(attachment.file_path):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attachment file not found on disk",
        )

    # For text/image/video/pdf files, show inline (in browser) instead of downloading
    inline_types = [
        "text/", "image/", "video/", "audio/",
        "application/pdf", "application/json", "application/xml",
    ]
    is_inline = any(attachment.mime_type and attachment.mime_type.startswith(t) for t in inline_types)

    if is_inline:
        return FileResponse(
            path=attachment.file_path,
            media_type=attachment.mime_type,
            headers={"Content-Disposition": f'inline; filename="{attachment.filename}"'},
        )

    return FileResponse(
        path=attachment.file_path,
        filename=attachment.filename,
        media_type=attachment.mime_type,
    )


@router.delete("/{ticket_id}/attachments/{attachment_id}")
async def delete_ticket_attachment(
    ticket_id: int,
    attachment_id: int,
    db: DbSession,
    current_user: Annotated[User, Depends(PermissionRequired("tickets.delete_attachments"))],
):
    """Delete an attachment."""
    result = await db.execute(
        select(TicketAttachment).where(
            TicketAttachment.id == attachment_id,
            TicketAttachment.ticket_id == ticket_id,
        )
    )
    attachment = result.scalar_one_or_none()

    if not attachment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Attachment not found",
        )

    # Delete file from disk
    if os.path.exists(attachment.file_path):
        os.remove(attachment.file_path)

    # Add history entry
    history = TicketHistory(
        ticket_id=ticket_id,
        user_id=current_user.id,
        action="attachment_deleted",
        old_value=json.dumps({"filename": attachment.filename}),
    )
    db.add(history)

    await db.delete(attachment)
    await db.commit()

    return {"message": "Attachment deleted successfully"}



